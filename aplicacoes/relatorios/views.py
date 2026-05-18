"""Views de relatórios — dashboard agregado, gráficos por categoria e exportações."""

import json
from io import BytesIO

from django.contrib.auth.decorators import login_required
from django.core.cache import cache
from django.db.models import Count
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render

from aplicacoes.adimplencia.models import Adimplencia
from aplicacoes.atividades.models import Atividade
from aplicacoes.cadastro.models import Municipio, Pessoa
from aplicacoes.engajamento.models import Engajamento
from aplicacoes.eventos.models import Evento, Participacao
from aplicacoes.instancias.models import Instancia, Representacao
from aplicacoes.missoes.models import Missao
from aplicacoes.projetos.models import Projeto


def _contagens_categorias():
    """Retorna contagens agregadas por categoria (com cache de 60s)."""
    chave = 'painel:contagens_categorias:v1'
    cacheado = cache.get(chave)
    if cacheado is not None:
        return cacheado
    resultado = _contagens_categorias_uncached()
    cache.set(chave, resultado, 60)
    return resultado


def _contagens_categorias_uncached():
    """Calcula contagens — extraído pra ficar fácil invalidar/refazer."""
    return {
        'instancias_origem': dict(
            Instancia.objects.values_list('origem')
            .annotate(total=Count('id')).values_list('origem', 'total')
        ),
        'projetos_status': dict(
            Projeto.objects.values_list('status')
            .annotate(total=Count('id')).values_list('status', 'total')
        ),
        'missoes_tipo': dict(
            Missao.objects.values_list('tipo')
            .annotate(total=Count('id')).values_list('tipo', 'total')
        ),
        'atividades_status': dict(
            Atividade.objects.values_list('status')
            .annotate(total=Count('id')).values_list('status', 'total')
        ),
    }


def _filtrar_municipios(queryset, regiao=None, uf=None):
    """Aplica filtros de região e UF a um queryset que tem FK para Municipio."""
    if regiao:
        queryset = queryset.filter(municipio__regiao=regiao)
    if uf:
        queryset = queryset.filter(municipio__uf=uf)
    return queryset


@login_required
def painel(request):
    """Painel de relatórios com KPIs, gráficos por categoria e top municípios.

    Suporta filtros via GET: ``regiao``, ``uf``. Aplicados sobre os dados
    sensíveis a município (engajamento, adimplência, top ranking).
    """
    regiao = request.GET.get('regiao', '')
    uf = request.GET.get('uf', '')

    total_pessoas = Pessoa.objects.filter(ativo=True).count()
    total_municipios = Municipio.objects.count()
    total_eventos = Evento.objects.count()
    total_participacoes = Participacao.objects.filter(confirmado=True).count()
    total_instancias = Instancia.objects.count()
    total_projetos = Projeto.objects.count()
    total_missoes = Missao.objects.count()
    total_atividades = Atividade.objects.count()
    total_representacoes = Representacao.objects.filter(vigente=True).count()

    ano_mais_recente = (
        Adimplencia.objects.order_by('-ano_referencia').values_list('ano_referencia', flat=True).first()
        or 2026
    )
    adimplencia_qs = Adimplencia.objects.filter(ano_referencia=ano_mais_recente)
    adimplencia_qs = _filtrar_municipios(adimplencia_qs, regiao, uf)
    adimplencia_status = dict(
        adimplencia_qs.values_list('status').annotate(total=Count('id')).values_list('status', 'total')
    )

    engajamento_qs = Engajamento.objects.all()
    engajamento_qs = _filtrar_municipios(engajamento_qs, regiao, uf)
    engajamento_niveis = dict(
        engajamento_qs.values_list('nivel').annotate(total=Count('id')).values_list('nivel', 'total')
    )

    municipios_regiao = dict(
        Municipio.objects.values_list('regiao')
        .annotate(total=Count('id')).values_list('regiao', 'total')
    )

    top_engajamento = engajamento_qs.select_related('municipio').order_by('-pontuacao_normalizada')[:10]

    contagens = _contagens_categorias()

    from aplicacoes.relatorios.servicos.narrativa import gerar_insights

    ctx = {
        # Narrativa do painel
        'insights': gerar_insights(),

        # Filtros
        'regiao_filtro': regiao, 'uf_filtro': uf,
        'regioes': Municipio.Regiao.choices,
        'ufs': Municipio.UF_CHOICES,

        # KPIs
        'total_pessoas': total_pessoas,
        'total_municipios': total_municipios,
        'total_eventos': total_eventos,
        'total_participacoes': total_participacoes,
        'total_instancias': total_instancias,
        'total_projetos': total_projetos,
        'total_missoes': total_missoes,
        'total_atividades': total_atividades,
        'total_representacoes': total_representacoes,
        'ano_adimplencia': ano_mais_recente,
        'top_engajamento': top_engajamento,

        # JSON para Chart.js
        'adimplencia_json': json.dumps({
            'labels': ['Adimplente', 'Inadimplente', 'Parcial'],
            'data': [
                adimplencia_status.get('adimplente', 0),
                adimplencia_status.get('inadimplente', 0),
                adimplencia_status.get('parcial', 0),
            ],
        }),
        'engajamento_json': json.dumps({
            'labels': ['Alto', 'Médio', 'Baixo', 'Inativo'],
            'data': [
                engajamento_niveis.get('alto', 0),
                engajamento_niveis.get('medio', 0),
                engajamento_niveis.get('baixo', 0),
                engajamento_niveis.get('inativo', 0),
            ],
        }),
        'regioes_json': json.dumps({
            'labels': ['Norte', 'Nordeste', 'Centro-Oeste', 'Sudeste', 'Sul'],
            'data': [
                municipios_regiao.get('norte', 0),
                municipios_regiao.get('nordeste', 0),
                municipios_regiao.get('centro_oeste', 0),
                municipios_regiao.get('sudeste', 0),
                municipios_regiao.get('sul', 0),
            ],
        }),
        'instancias_json': json.dumps({
            'labels': ['Internas', 'Externas', 'Eventos'],
            'data': [
                contagens['instancias_origem'].get('interna', 0),
                contagens['instancias_origem'].get('externa', 0),
                contagens['instancias_origem'].get('evento', 0),
            ],
        }),
        'projetos_json': json.dumps({
            'labels': ['Planejamento', 'Em andamento', 'Concluído', 'Pausado', 'Cancelado'],
            'data': [
                contagens['projetos_status'].get('planejamento', 0),
                contagens['projetos_status'].get('em_andamento', 0),
                contagens['projetos_status'].get('concluido', 0),
                contagens['projetos_status'].get('pausado', 0),
                contagens['projetos_status'].get('cancelado', 0),
            ],
        }),
        'missoes_json': json.dumps({
            'labels': ['Nacionais', 'Internacionais'],
            'data': [
                contagens['missoes_tipo'].get('nacional', 0),
                contagens['missoes_tipo'].get('internacional', 0),
            ],
        }),
        'atividades_json': json.dumps({
            'labels': ['Agendadas', 'Realizadas', 'Adiadas', 'Canceladas'],
            'data': [
                contagens['atividades_status'].get('agendada', 0),
                contagens['atividades_status'].get('realizada', 0),
                contagens['atividades_status'].get('adiada', 0),
                contagens['atividades_status'].get('cancelada', 0),
            ],
        }),
    }
    return render(request, 'relatorios/painel.html', ctx)


@login_required
def exportar_excel(request):
    """Exporta um workbook Excel com várias abas: engajamento, adimplência, pessoas e categorias."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')

    def _escreve_cabecalho(ws, headers):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    ws = wb.active
    ws.title = 'Engajamento'
    _escreve_cabecalho(ws, ['Município', 'UF', 'Região', 'Biênio', 'Pts Brutos', 'Pts Normalizado', 'Participações', 'Nível'])
    for row, eng in enumerate(Engajamento.objects.select_related('municipio').order_by('-pontuacao_bruta'), 2):
        ws.cell(row=row, column=1, value=eng.municipio.nome)
        ws.cell(row=row, column=2, value=eng.municipio.uf)
        ws.cell(row=row, column=3, value=eng.municipio.get_regiao_display())
        ws.cell(row=row, column=4, value=eng.bienio)
        ws.cell(row=row, column=5, value=eng.pontuacao_bruta)
        ws.cell(row=row, column=6, value=eng.pontuacao_normalizada)
        ws.cell(row=row, column=7, value=eng.total_participacoes)
        ws.cell(row=row, column=8, value=eng.get_nivel_display())

    ws2 = wb.create_sheet('Adimplência')
    _escreve_cabecalho(ws2, ['Município', 'UF', 'Ano', 'Status', 'Valor Devido', 'Valor Pago'])
    for row, a in enumerate(Adimplencia.objects.select_related('municipio').order_by('-ano_referencia', 'municipio__nome'), 2):
        ws2.cell(row=row, column=1, value=a.municipio.nome)
        ws2.cell(row=row, column=2, value=a.municipio.uf)
        ws2.cell(row=row, column=3, value=a.ano_referencia)
        ws2.cell(row=row, column=4, value=a.get_status_display())
        ws2.cell(row=row, column=5, value=float(a.valor_devido))
        ws2.cell(row=row, column=6, value=float(a.valor_pago))

    ws3 = wb.create_sheet('Pessoas')
    _escreve_cabecalho(ws3, ['Nome', 'Tipo', 'Cargo', 'Partido', 'E-mail', 'Status'])
    for row, p in enumerate(Pessoa.objects.filter(ativo=True).order_by('nome'), 2):
        ws3.cell(row=row, column=1, value=p.nome)
        ws3.cell(row=row, column=2, value=p.get_tipo_display())
        ws3.cell(row=row, column=3, value=p.cargo)
        ws3.cell(row=row, column=4, value=p.partido)
        ws3.cell(row=row, column=5, value=p.email or '')
        ws3.cell(row=row, column=6, value='Ativo' if p.ativo else 'Inativo')

    ws4 = wb.create_sheet('Instâncias')
    _escreve_cabecalho(ws4, ['Nome', 'Origem', 'Forma', 'Categoria', 'Status', 'Periodicidade', 'Ponto Focal'])
    for row, i in enumerate(Instancia.objects.select_related('ponto_focal_fnp').order_by('nome'), 2):
        ws4.cell(row=row, column=1, value=i.nome)
        ws4.cell(row=row, column=2, value=i.get_origem_display())
        ws4.cell(row=row, column=3, value=i.get_forma_display())
        ws4.cell(row=row, column=4, value=i.get_categoria_display())
        ws4.cell(row=row, column=5, value=i.get_status_display())
        ws4.cell(row=row, column=6, value=i.get_periodicidade_reunioes_display() if i.periodicidade_reunioes else '')
        ws4.cell(row=row, column=7, value=str(i.ponto_focal_fnp) if i.ponto_focal_fnp else '')

    ws5 = wb.create_sheet('Projetos')
    _escreve_cabecalho(ws5, ['Nome', 'Status', 'Fonte de recurso', 'Valor orçado', 'Início', 'Término previsto', 'Responsável'])
    for row, p in enumerate(Projeto.objects.select_related('responsavel').order_by('-data_inicio'), 2):
        ws5.cell(row=row, column=1, value=p.nome)
        ws5.cell(row=row, column=2, value=p.get_status_display())
        ws5.cell(row=row, column=3, value=p.get_fonte_recurso_display() if p.fonte_recurso else '')
        ws5.cell(row=row, column=4, value=float(p.valor_orcado) if p.valor_orcado else 0)
        ws5.cell(row=row, column=5, value=p.data_inicio)
        ws5.cell(row=row, column=6, value=p.data_fim_previsto)
        ws5.cell(row=row, column=7, value=str(p.responsavel) if p.responsavel else '')

    ws6 = wb.create_sheet('Missões')
    _escreve_cabecalho(ws6, ['Título', 'Tipo', 'País', 'Cidade', 'Status', 'Início', 'Término', 'Chefe'])
    for row, m in enumerate(Missao.objects.select_related('chefe_delegacao').order_by('-data_inicio'), 2):
        ws6.cell(row=row, column=1, value=m.titulo)
        ws6.cell(row=row, column=2, value=m.get_tipo_display())
        ws6.cell(row=row, column=3, value=m.pais)
        ws6.cell(row=row, column=4, value=m.cidade)
        ws6.cell(row=row, column=5, value=m.get_status_display())
        ws6.cell(row=row, column=6, value=m.data_inicio)
        ws6.cell(row=row, column=7, value=m.data_fim)
        ws6.cell(row=row, column=8, value=str(m.chefe_delegacao) if m.chefe_delegacao else '')

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="relatorio-fnp.xlsx"'
    return response


# Coordenadas aproximadas das capitais de UF — fallback para municipios sem lat/lng.
CAPITAIS_UF_COORD = {
    'AC': (-9.9747, -67.8243), 'AL': (-9.6498, -35.7089), 'AP': (0.0349, -51.0694),
    'AM': (-3.1190, -60.0217), 'BA': (-12.9714, -38.5014), 'CE': (-3.7172, -38.5433),
    'DF': (-15.7801, -47.9292), 'ES': (-20.3155, -40.3128), 'GO': (-16.6864, -49.2643),
    'MA': (-2.5307, -44.3068), 'MT': (-15.6010, -56.0974), 'MS': (-20.4486, -54.6295),
    'MG': (-19.9167, -43.9345), 'PA': (-1.4558, -48.4902), 'PB': (-7.1195, -34.8450),
    'PR': (-25.4284, -49.2733), 'PE': (-8.0476, -34.8770), 'PI': (-5.0892, -42.8019),
    'RJ': (-22.9068, -43.1729), 'RN': (-5.7945, -35.2110), 'RS': (-30.0346, -51.2177),
    'RO': (-8.7619, -63.9039), 'RR': (2.8235, -60.6758), 'SC': (-27.5954, -48.5480),
    'SP': (-23.5505, -46.6333), 'SE': (-10.9472, -37.0731), 'TO': (-10.1845, -48.3336),
}


def calendario_ics(request):
    """Feed iCal (.ics) com todos os eventos, atividades e missoes da FNP.

    Sem login_required para permitir subscribe direto no Google Calendar/Outlook
    via URL publica. Em prod, ativar acesso via token unico por usuario.
    """
    from datetime import datetime

    from django.http import HttpResponse

    from aplicacoes.atividades.models import Atividade
    from aplicacoes.eventos.models import Evento
    from aplicacoes.missoes.models import Missao

    linhas = ['BEGIN:VCALENDAR', 'VERSION:2.0', 'PRODID:-//FNP//Sistema FNP//PT-BR', 'X-WR-CALNAME:Agenda FNP', 'X-WR-TIMEZONE:America/Sao_Paulo']

    def _fmt(dt):
        if hasattr(dt, 'hour'):
            return dt.strftime('%Y%m%dT%H%M%S')
        return dt.strftime('%Y%m%d')

    def _ical_text(s):
        """Escapa caracteres especiais do iCal."""
        return (s or '').replace('\\', '\\\\').replace(',', '\\,').replace(';', '\\;').replace('\n', '\\n')

    agora = datetime.now()
    timestamp = agora.strftime('%Y%m%dT%H%M%SZ')

    # Eventos
    for ev in Evento.objects.all()[:500]:
        linhas.extend([
            'BEGIN:VEVENT',
            f'UID:evento-{ev.pk}@sistema-fnp',
            f'DTSTAMP:{timestamp}',
            f'DTSTART;VALUE=DATE:{_fmt(ev.data_inicio)}',
            f'DTEND;VALUE=DATE:{_fmt(ev.data_fim or ev.data_inicio)}',
            f'SUMMARY:{_ical_text(ev.titulo)}',
            f'DESCRIPTION:{_ical_text(ev.get_tipo_display())} · {_ical_text(ev.descricao or "")}',
            f'LOCATION:{_ical_text((ev.local or "") + " " + (ev.cidade or ""))}',
            'CATEGORIES:Evento FNP',
            'END:VEVENT',
        ])

    # Atividades
    for atv in Atividade.objects.select_related('instancia')[:500]:
        if atv.horario:
            dtstart = datetime.combine(atv.data_reuniao, atv.horario)
            linhas.extend([
                'BEGIN:VEVENT',
                f'UID:atividade-{atv.pk}@sistema-fnp',
                f'DTSTAMP:{timestamp}',
                f'DTSTART:{_fmt(dtstart)}',
                f'SUMMARY:{_ical_text(atv.titulo or atv.instancia.nome)}',
                f'DESCRIPTION:{_ical_text(atv.instancia.nome)} · {_ical_text(atv.get_tipo_calendario_display())}',
                f'LOCATION:{_ical_text(atv.local or "")}',
                'CATEGORIES:Atividade de Instância',
                'END:VEVENT',
            ])
        else:
            linhas.extend([
                'BEGIN:VEVENT',
                f'UID:atividade-{atv.pk}@sistema-fnp',
                f'DTSTAMP:{timestamp}',
                f'DTSTART;VALUE=DATE:{_fmt(atv.data_reuniao)}',
                f'SUMMARY:{_ical_text(atv.titulo or atv.instancia.nome)}',
                'CATEGORIES:Atividade de Instância',
                'END:VEVENT',
            ])

    # Missões
    for miss in Missao.objects.all()[:300]:
        linhas.extend([
            'BEGIN:VEVENT',
            f'UID:missao-{miss.pk}@sistema-fnp',
            f'DTSTAMP:{timestamp}',
            f'DTSTART;VALUE=DATE:{_fmt(miss.data_inicio)}',
            f'DTEND;VALUE=DATE:{_fmt(miss.data_fim or miss.data_inicio)}',
            f'SUMMARY:Missão: {_ical_text(miss.titulo)}',
            f'LOCATION:{_ical_text((miss.cidade or "") + ", " + (miss.pais or ""))}',
            'CATEGORIES:Missão',
            'END:VEVENT',
        ])

    linhas.append('END:VCALENDAR')

    response = HttpResponse('\r\n'.join(linhas), content_type='text/calendar; charset=utf-8')
    response['Content-Disposition'] = 'inline; filename="agenda-fnp.ics"'
    return response


@login_required
def calendario(request):
    """Pagina visual do calendario com FullCalendar.js (4 visoes).

    Renderiza um shell HTML e busca eventos via /api/calendario/eventos/.
    """
    return render(request, 'relatorios/calendario.html')


@login_required
def calendario_eventos_json(request):
    """JSON consumido pelo FullCalendar — eventos + atividades + missoes."""
    from django.http import JsonResponse

    from aplicacoes.atividades.models import Atividade
    from aplicacoes.eventos.models import Evento
    from aplicacoes.missoes.models import Missao

    start = request.GET.get('start')
    end = request.GET.get('end')

    eventos_json = []

    qs_ev = Evento.objects.all()
    if start and end:
        qs_ev = qs_ev.filter(data_inicio__gte=start[:10], data_inicio__lte=end[:10])
    for ev in qs_ev[:500]:
        eventos_json.append({
            'id': f'ev-{ev.pk}',
            'title': ev.titulo,
            'start': ev.data_inicio.isoformat(),
            'end': (ev.data_fim or ev.data_inicio).isoformat(),
            'url': f'/eventos/{ev.pk}/',
            'backgroundColor': '#3b82f6', 'borderColor': '#2563eb',
            'extendedProps': {'tipo': 'Evento', 'subtipo': ev.get_tipo_display(), 'local': ev.local or ''},
        })

    qs_atv = Atividade.objects.select_related('instancia')
    if start and end:
        qs_atv = qs_atv.filter(data_reuniao__gte=start[:10], data_reuniao__lte=end[:10])
    for atv in qs_atv[:500]:
        start_iso = atv.data_reuniao.isoformat()
        if atv.horario:
            start_iso = f'{atv.data_reuniao.isoformat()}T{atv.horario.isoformat()}'
        eventos_json.append({
            'id': f'atv-{atv.pk}',
            'title': atv.titulo or atv.instancia.nome,
            'start': start_iso,
            'url': f'/atividades/{atv.pk}/',
            'backgroundColor': '#14b8a6', 'borderColor': '#0d9488',
            'extendedProps': {'tipo': 'Atividade', 'subtipo': atv.instancia.nome, 'local': atv.local or ''},
        })

    qs_miss = Missao.objects.all()
    if start and end:
        qs_miss = qs_miss.filter(data_inicio__gte=start[:10], data_inicio__lte=end[:10])
    for miss in qs_miss[:300]:
        eventos_json.append({
            'id': f'miss-{miss.pk}',
            'title': f'✈️ {miss.titulo}',
            'start': miss.data_inicio.isoformat(),
            'end': (miss.data_fim or miss.data_inicio).isoformat(),
            'url': f'/missoes/{miss.pk}/',
            'backgroundColor': '#a855f7', 'borderColor': '#9333ea',
            'extendedProps': {'tipo': 'Missão', 'subtipo': miss.get_tipo_display(), 'local': f'{miss.cidade or ""}, {miss.pais or ""}'},
        })

    return JsonResponse(eventos_json, safe=False)


@login_required
def heatmap_pautas(request):
    """Cruza Pauta x Regiao para mostrar onde cada agenda tem mais envolvidos.

    Visualiza atraves de chart.js (heatmap-like via grid colorido). Insumo
    estrategico: "Norte foca seguranca, Sul mobilidade" -> direciona campanhas.
    """
    from django.db.models import Count

    from aplicacoes.cadastro.models import EnvolvimentoPauta, Municipio, Pauta

    pautas = list(Pauta.objects.filter(ativa=True).order_by('nome'))
    regioes_choices = Municipio.Regiao.choices
    regioes_slugs = [r[0] for r in regioes_choices]

    # Matriz [pauta][regiao] = contagem de pessoas com vinculo a municipio dessa regiao
    matriz = {p.id: {r: 0 for r in regioes_slugs} for p in pautas}
    envolvimentos = (
        EnvolvimentoPauta.objects
        .select_related('pessoa', 'pauta')
        .prefetch_related('pessoa__vinculos__municipio')
    )
    for env in envolvimentos:
        regioes_pessoa = set()
        for v in env.pessoa.vinculos.filter(vigente=True).select_related('municipio'):
            if v.municipio and v.municipio.regiao:
                regioes_pessoa.add(v.municipio.regiao)
        for r in regioes_pessoa:
            if env.pauta_id in matriz and r in matriz[env.pauta_id]:
                matriz[env.pauta_id][r] += 1

    # Achata para tabela renderizavel
    linhas = []
    for p in pautas:
        valores = [matriz[p.id][r] for r in regioes_slugs]
        max_val = max(valores) if valores else 0
        linhas.append({
            'pauta': p,
            'celulas': [
                {'regiao_label': dict(regioes_choices)[r], 'valor': matriz[p.id][r],
                 'intensidade': (matriz[p.id][r] / max_val * 100) if max_val else 0}
                for r in regioes_slugs
            ],
            'total': sum(valores),
        })
    linhas.sort(key=lambda x: x['total'], reverse=True)

    return render(request, 'relatorios/heatmap_pautas.html', {
        'linhas': linhas,
        'regioes_labels': [dict(regioes_choices)[r] for r in regioes_slugs],
    })


@login_required
def equipe_interna(request):
    """Visão de produtividade da equipe FNP interna.

    Mostra quem da equipe (Pessoa.tipo='interno') mais participou de eventos,
    fez missoes, marcou presenca etc. Util para gestao executiva da FNP.
    """
    from django.db.models import Count

    from aplicacoes.cadastro.models import Pessoa
    from aplicacoes.eventos.models import Participacao
    from aplicacoes.missoes.models import MembroDelegacao

    internos = Pessoa.objects.filter(tipo=Pessoa.TipoPessoa.INTERNO, ativo=True).order_by('nome')

    linhas = []
    for p in internos:
        n_part = Participacao.objects.filter(pessoa=p, confirmado=True).count()
        n_miss = MembroDelegacao.objects.filter(pessoa=p).count()
        n_pres = p.presencas.filter(status='presente').count() if hasattr(p, 'presencas') else 0
        n_repr = p.representacoes.filter(vigente=True).count() if hasattr(p, 'representacoes') else 0
        score = n_part + (n_miss * 3) + n_pres + (n_repr * 2)
        linhas.append({
            'pessoa': p,
            'participacoes': n_part,
            'missoes': n_miss,
            'presencas': n_pres,
            'representacoes': n_repr,
            'score': score,
        })
    linhas.sort(key=lambda x: x['score'], reverse=True)

    total_eventos_internos = sum(l['participacoes'] for l in linhas)
    total_missoes_internas = sum(l['missoes'] for l in linhas)

    return render(request, 'relatorios/equipe_interna.html', {
        'linhas': linhas,
        'total_internos': len(linhas),
        'total_eventos': total_eventos_internos,
        'total_missoes': total_missoes_internas,
    })


@login_required
def mapa(request):
    """Página de mapa do Brasil com choropleth UF + marcadores + heatmap.

    Suporta filtro temporal — usuário escolhe ano de referência e o mapa
    redesenha com snapshot daquele momento.
    """
    anos = list(
        Adimplencia.objects.values_list('ano_referencia', flat=True).distinct().order_by('-ano_referencia')
    )
    if not anos:
        anos = [2026]
    return render(request, 'relatorios/mapa.html', {
        'ano_atual': anos[0],
        'anos_disponiveis': anos,
    })


@login_required
def mapa_dados(request):
    """Endpoint JSON com pontos do mapa.

    Aceita ``modo=adimplencia`` (default) ou ``modo=engajamento`` via GET.
    Retorna apenas municípios com lat/lng ou que tenham status registrado,
    usando coordenada da capital da UF como fallback quando faltar geoposição.
    """
    modo = request.GET.get('modo', 'adimplencia')
    pontos = []

    if modo == 'engajamento':
        engajamentos = Engajamento.objects.select_related('municipio').all()
        # Mantém só o mais recente por município
        vistos = {}
        for eng in engajamentos.order_by('-bienio'):
            if eng.municipio_id in vistos:
                continue
            vistos[eng.municipio_id] = eng
        for eng in vistos.values():
            m = eng.municipio
            lat, lng = _coord_municipio(m)
            if lat is None:
                continue
            pontos.append({
                'lat': lat, 'lng': lng,
                'nome': m.nome, 'uf': m.uf,
                'status': eng.nivel,
                'rotulo': eng.get_nivel_display(),
                'valor': eng.pontuacao_normalizada,
                'id': str(m.id),
            })
    else:
        ano_req = request.GET.get('ano')
        ano = int(ano_req) if ano_req and ano_req.isdigit() else (
            Adimplencia.objects.order_by('-ano_referencia')
            .values_list('ano_referencia', flat=True).first() or 2026
        )
        adimplencias = (
            Adimplencia.objects
            .filter(ano_referencia=ano)
            .select_related('municipio')
        )
        for ad in adimplencias:
            m = ad.municipio
            lat, lng = _coord_municipio(m)
            if lat is None:
                continue
            pontos.append({
                'lat': lat, 'lng': lng,
                'nome': m.nome, 'uf': m.uf,
                'status': ad.status,
                'rotulo': ad.get_status_display(),
                'valor': float(ad.valor_pago) if ad.valor_pago else 0,
                'id': str(m.id),
            })

    return JsonResponse({'modo': modo, 'pontos': pontos, 'total': len(pontos)})


@login_required
def mapa_dados_uf(request):
    """Agregado por UF para a camada choropleth do mapa.

    Para ``modo=adimplencia`` retorna percentual de adimplentes no ano corrente.
    Para ``modo=engajamento`` retorna média da pontuação normalizada.
    """
    modo = request.GET.get('modo', 'adimplencia')
    agregados = {}

    if modo == 'engajamento':
        from django.db.models import Avg
        rows = (
            Engajamento.objects
            .values('municipio__uf')
            .annotate(media=Avg('pontuacao_normalizada'), total=Count('id'))
        )
        for r in rows:
            uf = r['municipio__uf']
            if not uf:
                continue
            agregados[uf] = {
                'valor': round(r['media'] or 0, 1),
                'total': r['total'],
                'rotulo': f"Engajamento médio: {round(r['media'] or 0, 1)}/100",
            }
    else:
        ano_req = request.GET.get('ano')
        ano = int(ano_req) if ano_req and ano_req.isdigit() else (
            Adimplencia.objects.order_by('-ano_referencia')
            .values_list('ano_referencia', flat=True).first() or 2026
        )
        rows = (
            Adimplencia.objects
            .filter(ano_referencia=ano)
            .values('municipio__uf', 'status')
            .annotate(total=Count('id'))
        )
        # Agrupa por UF e soma adimplentes vs total
        por_uf = {}
        for r in rows:
            uf = r['municipio__uf']
            if not uf:
                continue
            por_uf.setdefault(uf, {'adimplentes': 0, 'total': 0})
            por_uf[uf]['total'] += r['total']
            if r['status'] == 'adimplente':
                por_uf[uf]['adimplentes'] += r['total']
        for uf, dados in por_uf.items():
            pct = (dados['adimplentes'] / dados['total'] * 100) if dados['total'] else 0
            agregados[uf] = {
                'valor': round(pct, 1),
                'total': dados['total'],
                'rotulo': f"{round(pct, 1)}% adimplentes ({dados['adimplentes']}/{dados['total']})",
            }

    return JsonResponse({'modo': modo, 'agregados': agregados})


def _coord_municipio(municipio):
    """Retorna (lat, lng) do município ou da capital da UF como fallback.

    Returns:
        Tupla (lat, lng) como floats, ou (None, None) se não houver UF mapeada.
    """
    if municipio.latitude is not None and municipio.longitude is not None:
        return float(municipio.latitude), float(municipio.longitude)
    coord = CAPITAIS_UF_COORD.get(municipio.uf)
    if coord:
        # Jitter pequeno para não empilhar marcadores na mesma capital
        import hashlib
        h = int(hashlib.md5(str(municipio.id).encode()).hexdigest()[:8], 16)
        jx = ((h % 1000) / 1000 - 0.5) * 0.6  # ±0.3 graus
        jy = (((h // 1000) % 1000) / 1000 - 0.5) * 0.6
        return coord[0] + jx, coord[1] + jy
    return None, None


@login_required
def comparar_municipios(request):
    """Comparativo lado a lado de até 3 municípios para análise rápida.

    Aceita ``ids`` como GET (múltiplos valores ou lista CSV) com UUIDs de
    Municipio. Carrega métricas relevantes — população, engajamento atual,
    adimplência do ano mais recente, vínculos ativos e contagens de
    participações/representações — para renderização em cards verticais.
    """
    ids_raw = request.GET.getlist('ids')
    # Permite tanto ?ids=uuid1&ids=uuid2 quanto ?ids=uuid1,uuid2
    ids = []
    for chunk in ids_raw:
        ids.extend([x.strip() for x in chunk.split(',') if x.strip()])
    ids = ids[:3]  # Máximo 3 colunas para caber na tela

    municipios = list(Municipio.objects.filter(pk__in=ids))
    # Preserva a ordem original pedida via GET
    municipios.sort(key=lambda m: ids.index(str(m.pk)) if str(m.pk) in ids else 99)

    ano_recente = (
        Adimplencia.objects.order_by('-ano_referencia').values_list('ano_referencia', flat=True).first()
        or 2026
    )

    cartoes = []
    for municipio in municipios:
        engajamento = Engajamento.objects.filter(municipio=municipio).order_by('-bienio').first()
        adimplencia = Adimplencia.objects.filter(municipio=municipio, ano_referencia=ano_recente).first()
        cartoes.append({
            'municipio': municipio,
            'engajamento': engajamento,
            'adimplencia': adimplencia,
            'vinculos_ativos': municipio.vinculos.filter(vigente=True).count() if hasattr(municipio, 'vinculos') else 0,
            'participacoes': Participacao.objects.filter(municipio=municipio, confirmado=True).count(),
            'representacoes_vigentes': Representacao.objects.filter(pessoa__vinculos__municipio=municipio, vigente=True).distinct().count(),
        })

    # Lista para o autocomplete do seletor — serializada como JSON pronto pro Alpine
    disponiveis = [
        {'id': str(m.id), 'nome': m.nome, 'uf': m.uf}
        for m in Municipio.objects.order_by('nome').only('id', 'nome', 'uf')[:300]
    ]

    ctx = {
        'cartoes': cartoes,
        'ids_selecionados_json': json.dumps(ids),
        'ano_recente': ano_recente,
        'municipios_disponiveis_json': json.dumps(disponiveis),
    }
    return render(request, 'relatorios/comparar_municipios.html', ctx)


@login_required
def exportar_pdf(request):
    """Exporta relatório executivo resumido em PDF."""
    engajamentos = Engajamento.objects.select_related('municipio').order_by('-pontuacao_bruta')[:20]
    ctx = {
        'engajamentos': engajamentos,
        'total_pessoas': Pessoa.objects.filter(ativo=True).count(),
        'total_municipios': Municipio.objects.count(),
        'total_eventos': Evento.objects.count(),
        'total_instancias': Instancia.objects.count(),
        'total_projetos': Projeto.objects.count(),
        'total_missoes': Missao.objects.count(),
        'total_atividades': Atividade.objects.count(),
    }
    html = render(request, 'relatorios/pdf_relatorio.html', ctx).content.decode('utf-8')

    from xhtml2pdf import pisa
    pdf_buffer = BytesIO()
    pisa.CreatePDF(html, dest=pdf_buffer)
    pdf_buffer.seek(0)

    response = HttpResponse(pdf_buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="relatorio-fnp.pdf"'
    return response
